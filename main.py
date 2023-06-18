'''
This program provides options to search or add to an existing contact database.
The user will be prompted to add a contact, search for a contact or quit. Note
when adding or searching for a contact the query is case sensitive.
'''
import openpyxl
 
# collects user's first name, last name & number
def collect_contact_info():
    first_name = input("Enter first name: ")
    last_name = input("Enter last name: ")
    phone_number = input("Enter phone number: ")
    return [first_name, last_name, phone_number]

# writes collected contact info to xlsx file
def write_to_xlsx(contacts):
    workbook = openpyxl.load_workbook("contacts.xlsx")
    sheet = workbook.active
    sheet.append(contacts)
    workbook.save("contacts.xlsx")

# Searches for existing contact
def search_contacts(last_name):
    workbook = openpyxl.load_workbook("contacts.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[1] == last_name:
            print("Contact found:")
            print("First Name:", row[0])
            print("Last Name:", row[1])
            print("Phone Number:", row[2])
            return
    print("Contact not found.")

# main method prompts User
def main():
    while True:
        print("1. Add contact")
        print("2. Search contact")
        print("3. Quit")
        choice = input("Enter your choice (1-3): ")
        
        if choice == "1":
            contact_info = collect_contact_info()
            write_to_xlsx(contact_info)
            print("Contact added successfully!")
        elif choice == "2":
            last_name = input("Enter last name to search: ")
            search_contacts(last_name)
        elif choice == "3":
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()